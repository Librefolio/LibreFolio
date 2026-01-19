#!/usr/bin/env python3
"""
i18n Audit Tool for LibreFolio

This script analyzes translation files and generates a comparison table
showing all translation keys across all languages, highlighting missing translations.

Usage:
    python scripts/i18n-audit.py [--format md|xlsx|both] [--output path]

Examples:
    python scripts/i18n-audit.py                    # Markdown to stdout
    python scripts/i18n-audit.py --format xlsx     # Excel file in current dir
    python scripts/i18n-audit.py --format both     # Both formats in current dir
    python scripts/i18n-audit.py -o ./reports/     # Output to specific directory
"""

import json
import sys
from pathlib import Path
from typing import Any

# Try to import required libraries
try:
    import pandas as pd
except ImportError:
    print("❌ pandas is required for this script.")
    print("   Install with: pipenv install pandas")
    sys.exit(1)

try:
    from tabulate import tabulate
except ImportError:
    print("❌ tabulate is required for this script.")
    print("   Install with: pipenv install tabulate --dev")
    sys.exit(1)

# Configuration
I18N_DIR = Path(__file__).parent.parent / "src" / "lib" / "i18n"
LANGUAGES = ["en", "it", "fr", "es"]  # Order matters for display


def flatten_dict(d: dict, parent_key: str = "", sep: str = ".") -> dict[str, Any]:
    """
    Flatten a nested dictionary into dot-notation keys.

    Example:
        {"common": {"loading": "Loading..."}} -> {"common.loading": "Loading..."}
    """
    items: list[tuple[str, Any]] = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def load_translations() -> dict[str, dict[str, str]]:
    """
    Load all translation files and return flattened dictionaries.

    Returns:
        Dict mapping language code to flattened translations
    """
    translations: dict[str, dict[str, str]] = {}

    for lang in LANGUAGES:
        file_path = I18N_DIR / f"{lang}.json"
        if file_path.exists():
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                translations[lang] = flatten_dict(data)
                print(f"✓ Loaded {lang}.json ({len(translations[lang])} keys)")
            except json.JSONDecodeError as e:
                print(f"❌ Error parsing {lang}.json: {e}")
                translations[lang] = {}
        else:
            print(f"⚠ File not found: {lang}.json")
            translations[lang] = {}

    return translations


def get_all_keys(translations: dict[str, dict[str, str]]) -> list[str]:
    """
    Get all unique keys across all languages, sorted.
    """
    all_keys: set[str] = set()
    for lang_data in translations.values():
        all_keys.update(lang_data.keys())
    return sorted(all_keys)


def extract_section(key: str) -> str:
    """
    Extract the top-level section from a dotted key.

    Example: "common.loading" -> "common"
    """
    return key.split(".")[0] if "." in key else key


def build_dataframe(translations: dict[str, dict[str, str]]) -> pd.DataFrame:
    """
    Build a DataFrame with all translations.

    Columns: Section, Key, en, it, fr, es, Status
    """
    all_keys = get_all_keys(translations)

    rows: list[dict[str, Any]] = []
    for key in all_keys:
        row: dict[str, Any] = {
            "Section": extract_section(key),
            "Key": key,
            }

        # Add each language value
        missing_count = 0
        for lang in LANGUAGES:
            value = translations.get(lang, {}).get(key)
            row[lang.upper()] = value if value else ""
            if not value:
                missing_count += 1

        # Status column
        if missing_count == 0:
            row["Status"] = "✅"
        elif missing_count == len(LANGUAGES):
            row["Status"] = "❌ MISSING ALL"
        else:
            missing_langs = [lang.upper() for lang in LANGUAGES
                             if not translations.get(lang, {}).get(key)]
            row["Status"] = f"⚠️ Missing: {', '.join(missing_langs)}"

        rows.append(row)

    return pd.DataFrame(rows)


def generate_summary(df: pd.DataFrame) -> str:
    """
    Generate a summary of translation coverage.
    """
    lines = [
        "## 📊 Translation Coverage Summary\n",
        f"**Total keys**: {len(df)}\n",
        ]

    # Per-language stats
    for lang in LANGUAGES:
        col = lang.upper()
        filled = df[col].apply(lambda x: bool(x)).sum()
        pct = (filled / len(df)) * 100 if len(df) > 0 else 0
        lines.append(f"- **{col}**: {filled}/{len(df)} ({pct:.1f}%)")

    # Missing translations
    incomplete = df[df["Status"] != "✅"]
    if len(incomplete) > 0:
        lines.append(f"\n**⚠️ Incomplete translations**: {len(incomplete)}\n")
    else:
        lines.append("\n**✅ All translations complete!**\n")

    return "\n".join(lines)


def generate_missing_report(df: pd.DataFrame) -> str:
    """
    Generate a report of missing translations using tabulate for nice formatting.
    """
    incomplete = df[df["Status"] != "✅"].copy()

    if len(incomplete) == 0:
        return "\n## ✅ No Missing Translations\n\nAll keys are translated in all languages.\n"

    lines = [
        "\n## ⚠️ Missing Translations\n",
        "The following keys need attention:\n",
        ]

    # Group by section
    for section in incomplete["Section"].unique():
        section_rows = incomplete[incomplete["Section"] == section]
        lines.append(f"\n### {section}\n")

        # Build table data for tabulate
        table_data = []
        for _, row in section_rows.iterrows():
            table_data.append([f"`{row['Key']}`", row['Status']])

        # Use tabulate for nice formatting
        table_str = tabulate(
            table_data,
            headers=["Key", "Status"],
            tablefmt="github"
        )
        lines.append(table_str)
        lines.append("")

    return "\n".join(lines)


def export_markdown(df: pd.DataFrame, output_path: Path | None, include_full_table: bool = True) -> None:
    """
    Export the translation table to Markdown format using tabulate.

    Args:
        df: DataFrame with translations
        output_path: Output file path, or None for stdout
        include_full_table: If False, only show summary and missing translations
    """
    summary = generate_summary(df)
    missing_report = generate_missing_report(df)

    # Prepare DataFrame for tabulate (truncate long values)
    display_df = df.copy()
    for col in display_df.columns:
        display_df[col] = display_df[col].apply(
            lambda x: (str(x)[:10] + "...") if len(str(x)) > 13 else str(x)
            )

    # Generate markdown table using tabulate
    table_md = tabulate(
        display_df,
        headers="keys",
        tablefmt="github",  # GitHub-flavored markdown
        showindex=False
        )

    # Build report parts
    parts = [
        "# LibreFolio i18n Audit Report\n",
        f"*Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}*\n",
        summary,
        missing_report,
    ]

    if include_full_table:
        parts.extend([
            "\n## 📋 Complete Translation Table\n",
            table_md,
        ])

    parts.append("")
    content = "\n".join(parts)

    # Write to file or stdout
    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"\n✅ Markdown report saved to: {output_path}")
    else:
        print(content)


def export_excel(df: pd.DataFrame, output_path: Path) -> None:
    """
    Export the translation table to Excel format with formatting.
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("❌ openpyxl is required for Excel export.")
        print("   Install with: pip install openpyxl")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create Excel writer with formatting
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Main sheet with all translations
        df.to_excel(writer, sheet_name="All Translations", index=False)

        # Incomplete translations sheet
        incomplete = df[df["Status"] != "✅"]
        if len(incomplete) > 0:
            incomplete.to_excel(writer, sheet_name="Missing", index=False)

        # Summary sheet
        summary_data = {
            "Language": [lang.upper() for lang in LANGUAGES],
            "Filled": [df[lang.upper()].apply(lambda x: bool(x)).sum() for lang in LANGUAGES],
            "Missing": [df[lang.upper()].apply(lambda x: not bool(x)).sum() for lang in LANGUAGES],
            "Coverage %": [
                f"{(df[lang.upper()].apply(lambda x: bool(x)).sum() / len(df)) * 100:.1f}%"
                for lang in LANGUAGES
                ],
            }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Apply formatting
        workbook = writer.book

        # Format header row
        header_fill = PatternFill(start_color="1A4D3E", end_color="1A4D3E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

    print(f"\n✅ Excel report saved to: {output_path}")


def run_audit(format_type: str = "none", output: str | None = None) -> int:
    """Run the i18n audit with specified format and output."""
    # Determine output directory (PWD by default)
    if output:
        output_base = Path(output)
        if output_base.suffix in [".md", ".xlsx"]:
            # User specified a file, use its parent as base
            output_dir = output_base.parent
        else:
            output_dir = output_base
    else:
        output_dir = Path.cwd()

    print("=" * 60)
    print("  LibreFolio i18n Audit Tool")
    print("=" * 60)
    print()

    # Load translations
    print("📂 Loading translation files...")
    translations = load_translations()
    print()

    # Build DataFrame
    print("🔍 Analyzing translations...")
    df = build_dataframe(translations)

    # Generate output based on format
    if format_type == "none":
        # Only show summary and warnings, no full table
        export_markdown(df, None, include_full_table=False)
    elif format_type in ["md", "both"]:
        if output and output.endswith(".md"):
            md_path = Path(output)
        elif format_type == "md" and output is None:
            # Print to stdout if no output specified for md-only
            export_markdown(df, None, include_full_table=True)
            md_path = None
        else:
            md_path = output_dir / "i18n-audit.md"

        if md_path:
            export_markdown(df, md_path, include_full_table=True)

    if format_type in ["xlsx", "both"]:
        if output and output.endswith(".xlsx"):
            xlsx_path = Path(output)
        else:
            xlsx_path = output_dir / "i18n-audit.xlsx"
        export_excel(df, xlsx_path)

    # Print summary
    print("\n" + "=" * 60)
    print("  Summary")
    print("=" * 60)
    print(f"  Total keys: {len(df)}")
    complete = len(df[df["Status"] == "✅"])
    incomplete = len(df) - complete
    print(f"  Complete:   {complete} ✅")
    print(f"  Incomplete: {incomplete} {'⚠️' if incomplete > 0 else ''}")
    print()

    return 0


def add_arguments(parser) -> None:
    """Add arguments to a parser (reusable for both standalone and subparser)."""
    parser.add_argument(
        "--format", "-f",
        choices=["none", "md", "xlsx", "both"],
        default="none",
        help="Output format: none=only warnings (default), md=full markdown, xlsx=excel, both=md+xlsx"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output directory or file path (default: current directory for file output)"
    )


def run_from_args(args) -> int:
    """Execute the command from parsed args."""
    return run_audit(
        format_type=getattr(args, 'format', 'none'),
        output=getattr(args, 'output', None)
    )


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Audit LibreFolio i18n translations",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python i18n-audit.py                    Show warnings only (no full table)
  python i18n-audit.py --format md        Show full Markdown report in terminal
  python i18n-audit.py --format xlsx      Generate Excel file in current dir
  python i18n-audit.py --format both      Generate both formats in current dir
  python i18n-audit.py -o ./reports/      Save to specific directory
        """
    )
    add_arguments(parser)
    args = parser.parse_args()
    return run_from_args(args)


def register_subparser(subparsers) -> None:
    """Register as subparser for dev.py integration."""
    p = subparsers.add_parser("i18n", help="📦 Translation commands")
    i18n_sub = p.add_subparsers(dest="i18n_cmd", metavar="action")

    audit_p = i18n_sub.add_parser("audit", help="Audit translations for missing keys")
    add_arguments(audit_p)
    audit_p.set_defaults(func=run_from_args)


if __name__ == "__main__":
    main()
